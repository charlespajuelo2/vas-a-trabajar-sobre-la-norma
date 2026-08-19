from __future__ import annotations

import math
import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

import openpyxl

try:  # Streamlit is required to run the UI, but tests can run without it.
    import streamlit as st
except ImportError:  # pragma: no cover - exercised only when testing without Streamlit installed.
    st = None


BASE_DIR = Path(__file__).resolve().parent
SOURCE_PATH = BASE_DIR / "E060_Acero_Usuario.xlsx"
PLATE_PATH = BASE_DIR / "outputs" / "lamina_detalle_acero_refuerzo.svg"
APP_NAME = "E060 Acero — Usuario"

TRACE_COLORS = {
    "NORMATIVO": "#15803D",
    "DERIVADO": "#D97706",
    "NO VERIFICADO": "#B91C1C",
}


def as_number(value: Any) -> Optional[float]:
    """Return a finite numeric value, or None for blank/text values."""
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)) and math.isfinite(float(value)):
        return float(value)
    return None


def nonblank(value: Any) -> bool:
    return value is not None and str(value).strip() != ""


def key_value(sources: Dict[str, Dict[str, Any]], key: str) -> Optional[float]:
    record = sources.get(key)
    if not record or record.get("state") != "VERIFICADO":
        return None
    return as_number(record.get("value"))


class ExcelSource:
    """Read-only adapter around the workbook. No normative value is duplicated here."""

    def __init__(self, path: Path):
        self.path = path
        if not path.exists():
            raise FileNotFoundError(f"No se encontró la única fuente Excel: {path}")
        self.formula_wb = openpyxl.load_workbook(path, data_only=False, read_only=False)
        self.value_wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
        self.sources = self._load_sources()
        self.bars = self._load_bars()
        self.cases = self._load_cases()
        self.calc_rows = self._load_calculations()
        self.entry_notes = self._load_entry_notes()

    def _load_sources(self) -> Dict[str, Dict[str, Any]]:
        ws = self.formula_wb["FUENTES"]
        data: Dict[str, Dict[str, Any]] = {}
        for row in range(5, ws.max_row + 1):
            key = ws.cell(row, 1).value
            if not nonblank(key):
                continue
            data[str(key)] = {
                "key": str(key),
                "layer": ws.cell(row, 2).value,
                "reference": ws.cell(row, 3).value,
                "page": ws.cell(row, 4).value,
                "state": ws.cell(row, 5).value,
                "value": ws.cell(row, 6).value,
                "unit": ws.cell(row, 7).value,
                "note": ws.cell(row, 8).value,
            }
        return data

    def _load_bars(self) -> List[Dict[str, Any]]:
        ws = self.formula_wb["BARRAS"]
        data: List[Dict[str, Any]] = []
        for row in range(5, ws.max_row + 1):
            label = ws.cell(row, 1).value
            if not nonblank(label):
                continue
            data.append(
                {
                    "row": row,
                    "commercial": label,
                    "db": ws.cell(row, 2).value,
                    "area": ws.cell(row, 3).value,
                    "perimeter": ws.cell(row, 4).value,
                    "db_formula": ws.cell(row, 2).value if isinstance(ws.cell(row, 2).value, str) and ws.cell(row, 2).value.startswith("=") else f"Dato directo de BARRAS!B{row}",
                    "area_formula": ws.cell(row, 3).value if isinstance(ws.cell(row, 3).value, str) and ws.cell(row, 3).value.startswith("=") else f"Dato directo de BARRAS!C{row}",
                    "perimeter_formula": ws.cell(row, 4).value if isinstance(ws.cell(row, 4).value, str) and ws.cell(row, 4).value.startswith("=") else f"Dato directo de BARRAS!D{row}",
                    "reference": ws.cell(row, 5).value,
                    "state": ws.cell(row, 6).value,
                }
            )
        return data

    def _load_cases(self) -> List[Dict[str, Any]]:
        ws = self.value_wb["ENTRADAS"]
        headers = {ws.cell(5, col).value: col for col in range(7, 25)}
        cases: List[Dict[str, Any]] = []
        for row in range(6, 12):
            case_id = ws.cell(row, headers["ID"]).value
            if not nonblank(case_id):
                continue
            cases.append(
                {
                    "id": case_id,
                    "name": ws.cell(row, headers["Caso"]).value,
                    "fc": ws.cell(row, headers["f'c"]).value,
                    "fy": ws.cell(row, headers["fy"]).value,
                    "db": ws.cell(row, headers["db"]).value,
                    "range": ws.cell(row, headers["Rango"]).value,
                    "cover": ws.cell(row, headers["Recub."]).value,
                    "spacing": ws.cell(row, headers["s c-c"]).value,
                    "psi_t": ws.cell(row, headers["ψt"]).value,
                    "psi_e": ws.cell(row, headers["ψe"]).value,
                    "psi_s": ws.cell(row, headers["ψs"]).value,
                    "lambda": ws.cell(row, headers["λ"]).value,
                    "ktr": ws.cell(row, headers["Ktr"]).value,
                    "as_req": ws.cell(row, headers["As req"]).value,
                    "as_prov": ws.cell(row, headers["As prov"]).value,
                    "pct_spliced": ws.cell(row, headers["% empalme"]).value,
                    "ld_base": ws.cell(row, headers["ld base"]).value,
                    "expected": ws.cell(row, headers["Esperado"]).value,
                }
            )
        return cases

    def _load_calculations(self) -> List[Dict[str, Any]]:
        ws = self.formula_wb["CALCULOS"]
        rows: List[Dict[str, Any]] = []
        for row in range(5, ws.max_row + 1):
            concept = ws.cell(row, 1).value
            if not nonblank(concept):
                continue
            rows.append(
                {
                    "row": row,
                    "concept": str(concept),
                    "formula": ws.cell(row, 2).value,
                    "unit": ws.cell(row, 3).value,
                    "layer": ws.cell(row, 4).value,
                    "reference": ws.cell(row, 5).value,
                    "dependency": ws.cell(row, 6).value,
                }
            )
        return rows

    def _load_entry_notes(self) -> Dict[str, str]:
        ws = self.formula_wb["ENTRADAS"]
        return {
            "bar_types": [part.strip() for part in str(ws.cell(7, 4).value or "").split("/") if part.strip()],
            "ranges": [part.strip() for part in str(ws.cell(23, 4).value or "").split(";") if part.strip()],
            "conditions": str(ws.cell(12, 4).value or ""),
        }

    def calc_meta(self, contains: str) -> Dict[str, Any]:
        needle = contains.casefold()
        for row in self.calc_rows:
            if needle in row["concept"].casefold():
                return row
        raise KeyError(f"No se encontró el cálculo en CALCULOS: {contains}")


def trace_result(
    source: ExcelSource,
    meta: Dict[str, Any],
    value: Any,
    status: str,
    dependency: Optional[str] = None,
) -> Dict[str, Any]:
    return {
        "concept": meta["concept"],
        "value": value,
        "status": status,
        "formula": meta["formula"],
        "unit": meta["unit"],
        "reference": meta["reference"],
        "dependency": dependency or meta["dependency"],
    }


def no_verified(source: ExcelSource, meta: Dict[str, Any], dependency: str) -> Dict[str, Any]:
    return trace_result(source, meta, "NO VERIFICADO", "NO VERIFICADO", dependency)


def calculate_hook(
    source: ExcelSource,
    hook_index: int,
    db: Any,
    bar_type: str,
    range_label: str,
) -> Dict[str, Any]:
    meta = [row for row in source.calc_rows if "extensión gancho" in row["concept"].casefold()][hook_index]
    diameter = as_number(db)
    if diameter is None:
        return no_verified(source, meta, "depende de db verificado")

    if hook_index == 0:
        k = key_value(source.sources, "k_ext_180")
        minimum = key_value(source.sources, "min_ext_180")
        if k is None or minimum is None:
            return no_verified(source, meta, "falta un parámetro FUENTES verificado")
        return trace_result(source, meta, max(diameter * k, minimum), "DERIVADO")
    bar_type_options = source.entry_notes["bar_types"]
    range_options = source.entry_notes["ranges"]
    longitudinal_type = bar_type_options[0] if bar_type_options else None
    stirrup_type = bar_type_options[1] if len(bar_type_options) > 1 else None
    small_range = range_options[0] if range_options else None
    large_range = range_options[1] if len(range_options) > 1 else None

    if hook_index == 1:
        if bar_type != longitudinal_type:
            return no_verified(source, meta, "requiere tipo longitudinal")
        k = key_value(source.sources, "k_ext_90_long")
        if k is None:
            return no_verified(source, meta, "falta k_ext_90_long")
        return trace_result(source, meta, diameter * k, "DERIVADO")
    if hook_index == 2:
        if bar_type != stirrup_type or range_label != small_range:
            return no_verified(source, meta, "requiere estribo-grapa y el primer rango de la fuente")
        k = key_value(source.sources, "k_ext_90_stirrup_small")
        if k is None:
            return no_verified(source, meta, "falta k_ext_90_stirrup_small")
        return trace_result(source, meta, diameter * k, "DERIVADO")
    if hook_index == 3:
        if bar_type != stirrup_type or range_label != large_range:
            return no_verified(source, meta, "requiere estribo-grapa y el segundo rango de la fuente")
        k = key_value(source.sources, "k_ext_90_stirrup_large")
        if k is None:
            return no_verified(source, meta, "falta k_ext_90_stirrup_large")
        return trace_result(source, meta, diameter * k, "DERIVADO")
    if bar_type != stirrup_type or range_label != small_range:
        return no_verified(source, meta, "requiere estribo-grapa y el rango aceptado por CALCULOS")
    k = key_value(source.sources, "k_ext_135")
    if k is None:
        return no_verified(source, meta, "falta k_ext_135")
    return trace_result(source, meta, diameter * k, "DERIVADO")


def calculate_bend(source: ExcelSource, db: Any, bar_type: str, range_label: str) -> Dict[str, Any]:
    meta = source.calc_meta("Diámetro interior mínimo de doblado")
    diameter = as_number(db)
    if diameter is None or not nonblank(range_label):
        return no_verified(source, meta, "depende de db y rango de doblado")
    range_options = source.entry_notes["ranges"]
    bar_type_options = source.entry_notes["bar_types"]
    small_range = range_options[0] if range_options else None
    first_table_range = range_options[2] if len(range_options) > 2 else None
    second_table_range = range_options[3] if len(range_options) > 3 else None
    third_table_range = range_options[4] if len(range_options) > 4 else None
    stirrup_type = bar_type_options[1] if len(bar_type_options) > 1 else None
    key = None
    if bar_type == stirrup_type and range_label == small_range:
        key = "k_bend_stirrup_small"
    elif range_label == first_table_range:
        key = "k_bend_1_4_1"
    elif range_label == second_table_range:
        key = "k_bend_1_1_8_1_3_8"
    elif range_label == third_table_range:
        key = "k_bend_1_11_16_2_1_4"
    if key is None:
        return no_verified(source, meta, "rango sin parámetro verificado en FUENTES")
    factor = key_value(source.sources, key)
    if factor is None:
        return no_verified(source, meta, f"falta {key}")
    return trace_result(source, meta, diameter * factor, "DERIVADO")


def calculate_development_and_laps(source: ExcelSource, values: Dict[str, Any]) -> Dict[str, Dict[str, Any]]:
    fc = as_number(values.get("fc"))
    fy = as_number(values.get("fy"))
    db = as_number(values.get("db"))
    cover = as_number(values.get("cover"))
    spacing = as_number(values.get("spacing"))
    psi_t = as_number(values.get("psi_t"))
    psi_e = as_number(values.get("psi_e"))
    psi_s = as_number(values.get("psi_s"))
    lambda_factor = as_number(values.get("lambda"))
    ktr = as_number(values.get("ktr"))

    def has_all(*items: Any) -> bool:
        return all(as_number(item) is not None for item in items)

    ld_t_meta = source.calc_meta("ld en tracción")
    if has_all(fc, fy, db, cover, spacing, psi_t, psi_e, psi_s, lambda_factor, ktr):
        fc_n, fy_n, db_n = float(fc), float(fy), float(db)
        half = key_value(source.sources, "half_divisor")
        coefficient = key_value(source.sources, "tension_formula_coef")
        denominator_cap = key_value(source.sources, "tension_den_cap")
        minimum = key_value(source.sources, "ld_tension_min")
        psi_cap = key_value(source.sources, "psi_te_cap")
        if None in (half, coefficient, denominator_cap, minimum, psi_cap) or fc_n <= 0 or db_n <= 0:
            ld_t = no_verified(source, ld_t_meta, "falta parámetro verificado o hay un denominador no válido")
        else:
            cb = min(float(cover) + db_n / half, float(spacing) / half)
            denominator = min(denominator_cap, (cb + float(ktr)) / db_n)
            if denominator <= 0 or float(lambda_factor) <= 0:
                ld_t = no_verified(source, ld_t_meta, "denominador de 12.2.3 no válido")
            else:
                base = (
                    float(fy)
                    * min(psi_cap, float(psi_t) * float(psi_e))
                    * float(psi_s)
                    * db_n
                    / (coefficient * float(lambda_factor) * math.sqrt(fc_n) * denominator)
                )
                ld_t = trace_result(source, ld_t_meta, max(minimum, base), "DERIVADO")
    else:
        ld_t = no_verified(source, ld_t_meta, "depende de f'c, fy, db, recubrimiento, espaciamiento, Ktr y ψ")

    ld_c_meta = source.calc_meta("ld en compresión")
    if has_all(fc, fy, db) and float(fc) > 0:
        a = key_value(source.sources, "ld_compression_coef_a")
        b = key_value(source.sources, "ld_compression_coef_b")
        minimum = key_value(source.sources, "ld_compression_min")
        if None in (a, b, minimum):
            ld_c = no_verified(source, ld_c_meta, "falta parámetro de 12.3.2")
        else:
            base_a = a * float(fy) / math.sqrt(float(fc)) * float(db)
            base_b = b * float(fy) * float(db)
            ld_c = trace_result(source, ld_c_meta, max(minimum, base_a, base_b), "DERIVADO")
    else:
        ld_c = no_verified(source, ld_c_meta, "depende de f'c, fy y db")

    ldg_meta = source.calc_meta("ldg de gancho")
    if has_all(fc, fy, db, psi_e, lambda_factor) and float(fc) > 0:
        coef = key_value(source.sources, "hook_ld_coef")
        min_db = key_value(source.sources, "ld_hook_min_db")
        min_mm = key_value(source.sources, "ld_hook_min_mm")
        if None in (coef, min_db, min_mm):
            ldg = no_verified(source, ldg_meta, "falta parámetro de 12.5")
        else:
            base = coef * float(psi_e) * float(lambda_factor) * float(fy) / math.sqrt(float(fc)) * float(db)
            ldg = trace_result(source, ldg_meta, max(min(min_db * float(db), min_mm), base), "DERIVADO")
    else:
        ldg = no_verified(source, ldg_meta, "depende de f'c, fy, db, ψe y λ")

    lap_t_meta = source.calc_meta("Empalme por traslape en tracción")
    as_req = as_number(values.get("as_req"))
    as_prov = as_number(values.get("as_prov"))
    pct = as_number(values.get("pct_spliced"))
    ld_base_override = as_number(values.get("ld_base"))
    if (ld_t["status"] == "DERIVADO" or ld_base_override is not None) and has_all(as_req, as_prov, pct) and as_req > 0:
        ratio_min = key_value(source.sources, "ratio_class_A_min")
        pct_max = key_value(source.sources, "pct_class_A_max")
        factor_a = key_value(source.sources, "lap_tension_class_A_factor")
        factor_b = key_value(source.sources, "lap_tension_class_B_factor")
        minimum = key_value(source.sources, "lap_min")
        if None in (ratio_min, pct_max, factor_a, factor_b, minimum):
            lap_t = no_verified(source, lap_t_meta, "falta parámetro de Tabla 12.3")
        else:
            factor = factor_a if as_prov / as_req >= ratio_min and pct <= pct_max else factor_b
            base_ld = ld_base_override if ld_base_override is not None else float(ld_t["value"])
            lap_t = trace_result(source, lap_t_meta, max(minimum, factor * base_ld), "DERIVADO")
    else:
        lap_t = no_verified(source, lap_t_meta, "depende de ld en tracción, As prov/As req y porcentaje empalmado")

    lap_c_meta = source.calc_meta("Empalme por traslape en compresión")
    if has_all(fc, fy, db) and float(fc) > 0:
        coef_le = key_value(source.sources, "lap_compression_coef_le_420")
        coef_gt = key_value(source.sources, "lap_compression_coef_gt_420")
        subtract = key_value(source.sources, "lap_compression_subtract")
        threshold_fy = key_value(source.sources, "fy_threshold_compression")
        threshold_fc = key_value(source.sources, "fc_threshold_compression")
        fc_factor = key_value(source.sources, "lap_compression_fc_factor")
        one = key_value(source.sources, "one")
        minimum = key_value(source.sources, "lap_min")
        if None in (coef_le, coef_gt, subtract, threshold_fy, threshold_fc, fc_factor, one, minimum):
            lap_c = no_verified(source, lap_c_meta, "falta parámetro de 12.16.1")
        else:
            core = coef_le * float(fy) * float(db) if float(fy) <= threshold_fy else (coef_gt * float(fy) - subtract) * float(db)
            factor = fc_factor if float(fc) < threshold_fc else one
            lap_c = trace_result(source, lap_c_meta, max(minimum, core * factor), "DERIVADO")
    else:
        lap_c = no_verified(source, lap_c_meta, "depende de f'c, fy y db")

    column_meta = source.calc_meta("Empalme de columna")
    confinement = as_number(values.get("confinement"))
    if lap_c["status"] == "DERIVADO" and confinement is not None:
        column = trace_result(source, column_meta, lap_c["value"] * confinement, "DERIVADO")
    else:
        column = no_verified(source, column_meta, "depende de empalme en compresión y factor de confinamiento")

    return {
        "ld_tension": ld_t,
        "ld_compression": ld_c,
        "ldg": ldg,
        "lap_tension": lap_t,
        "lap_compression": lap_c,
        "column_splice": column,
    }


def _split_note(note: str, separator: str) -> List[str]:
    return [part.strip() for part in note.split(separator) if part.strip()]


def status_badge(status: str) -> str:
    color = TRACE_COLORS.get(status, TRACE_COLORS["NO VERIFICADO"])
    return f'<span style="color:{color};font-weight:700">{status}</span>'


def render_trace_card(result: Dict[str, Any], source: ExcelSource) -> None:
    if st is None:
        return
    status = result["status"]
    value = result["value"]
    with st.container(border=True):
        st.markdown(f"**{result['concept']}**")
        if status == "NO VERIFICADO":
            st.error("NO VERIFICADO")
        else:
            shown = f"{value:,.2f} {result['unit']}" if isinstance(value, (int, float)) else str(value)
            st.metric("Resultado", shown)
            st.markdown(f"Semáforo: {status_badge(status)}", unsafe_allow_html=True)
        st.markdown(f"**Fórmula utilizada:** `{result['formula']}`")
        st.markdown(f"**Unidades:** {result['unit']}  \n**Referencia normativa:** {result['reference']}")
        st.caption(result["dependency"])


def render_source_card(label: str, value: Any, unit: Any, reference: Any, state: Any, formula: Any) -> None:
    status = state if state in TRACE_COLORS else "NO VERIFICADO"
    with st.container(border=True):
        st.markdown(f"**{label}**")
        if status == "NO VERIFICADO":
            st.error("NO VERIFICADO")
        else:
            st.metric("Valor", f"{value} {unit or ''}")
            st.markdown(f"Semáforo: {status_badge(status)}", unsafe_allow_html=True)
        st.markdown(f"**Fórmula utilizada:** `{formula}`")
        st.markdown(f"**Unidades:** {unit or '—'}")
        st.markdown(f"**Referencia normativa:** {reference}")


def main() -> None:
    if st is None:
        raise RuntimeError("Instala las dependencias de requirements.txt para ejecutar Streamlit.")

    st.set_page_config(page_title=APP_NAME, layout="wide")
    st.markdown(
        """
        <style>
        .fixed-warning {position: sticky; top: 0; z-index: 1000; background: #7f1d1d; color: white;
                        padding: 0.8rem 1rem; border-radius: 0.35rem; font-weight: 700; margin-bottom: 1rem;}
        .source-caption {color: #475569; font-size: 0.9rem;}
        </style>
        <div class="fixed-warning">Herramienta educativa. No apta para diseño estructural.<br>
        Los valores deben verificarse contra la Norma E.060 vigente.</div>
        """,
        unsafe_allow_html=True,
    )

    st.title(APP_NAME)
    st.caption(f"Fuente única: {SOURCE_PATH.name}")
    try:
        source = ExcelSource(SOURCE_PATH)
    except Exception as exc:
        st.error(str(exc))
        st.stop()

    hooks = [row for row in source.calc_rows if "extensión gancho" in row["concept"].casefold()]
    hook_labels = [row["concept"] for row in hooks]
    bar_types = source.entry_notes["bar_types"]
    ranges = source.entry_notes["ranges"]
    cases = source.cases
    case_labels = [f"{case['id']} — {case['name']}" for case in cases]

    controls, details = st.tabs(["Cálculo trazable", "Lámina técnica"])
    with controls:
        st.subheader("Entradas desde el Excel")
        left, right = st.columns(2)
        with left:
            selected_bar_label = st.selectbox("Diámetro comercial", [bar["commercial"] for bar in source.bars])
            selected_bar = next(bar for bar in source.bars if bar["commercial"] == selected_bar_label)
            selected_case_label = st.selectbox("Caso de prueba de VALIDACION (editable como punto de partida)", case_labels)
            case = cases[case_labels.index(selected_case_label)]
            default_db = as_number(case["db"])
            db = st.number_input("db usado en el cálculo (mm)", min_value=key_value(source.sources, "zero") or 0.0, value=default_db or 0.0, step=key_value(source.sources, "one") or 1.0)
            bar_type_default = bar_types[1] if len(bar_types) > 1 and "estribo" in str(case["name"]).casefold() else bar_types[0]
            bar_type = st.selectbox("Tipo de barra", bar_types, index=bar_types.index(bar_type_default) if bar_type_default in bar_types else 0)
            range_default = str(case["range"])
            range_label = st.selectbox("Clase/rango para gancho y doblado", ranges, index=ranges.index(range_default) if range_default in ranges else 0)
            hook_index = st.selectbox("Tipo de gancho", list(range(len(hook_labels))), format_func=lambda idx: hook_labels[idx])
        with right:
            fc = st.number_input("f'c (MPa)", min_value=key_value(source.sources, "zero") or 0.0, value=float(case["fc"]), step=key_value(source.sources, "one") or 1.0)
            fy = st.number_input("fy (MPa)", min_value=key_value(source.sources, "zero") or 0.0, value=float(case["fy"]), step=key_value(source.sources, "one") or 1.0)
            cover = st.number_input("Recubrimiento claro (mm)", min_value=key_value(source.sources, "zero") or 0.0, value=float(case["cover"]), step=key_value(source.sources, "one") or 1.0)
            spacing = st.number_input("Espaciamiento centro a centro (mm)", min_value=key_value(source.sources, "zero") or 0.0, value=float(case["spacing"]), step=key_value(source.sources, "one") or 1.0)
            ktr = st.number_input("Ktr (mm)", min_value=key_value(source.sources, "zero") or 0.0, value=float(case["ktr"]), step=key_value(source.sources, "one") or 1.0)

        with st.expander("Factores y datos para desarrollo/traslape", expanded=False):
            c1, c2, c3 = st.columns(3)
            with c1:
                psi_t = st.number_input("ψt", min_value=key_value(source.sources, "zero") or 0.0, value=float(case["psi_t"]), step=key_value(source.sources, "one") or 1.0)
                psi_e = st.number_input("ψe", min_value=key_value(source.sources, "zero") or 0.0, value=float(case["psi_e"]), step=key_value(source.sources, "one") or 1.0)
            with c2:
                psi_s = st.number_input("ψs", min_value=key_value(source.sources, "zero") or 0.0, value=float(case["psi_s"]), step=key_value(source.sources, "one") or 1.0)
                lambda_factor = st.number_input("λ", min_value=key_value(source.sources, "zero") or 0.0, value=float(case["lambda"]), step=key_value(source.sources, "one") or 1.0)
            with c3:
                as_req = st.number_input("As requerida (mm²)", min_value=key_value(source.sources, "zero") or 0.0, value=float(case["as_req"]), step=key_value(source.sources, "one") or 1.0)
                as_prov = st.number_input("As provista (mm²)", min_value=key_value(source.sources, "zero") or 0.0, value=float(case["as_prov"]), step=key_value(source.sources, "one") or 1.0)
                pct_spliced = st.number_input("Porcentaje de barras empalmadas (fracción)", min_value=key_value(source.sources, "zero") or 0.0, max_value=key_value(source.sources, "one") or 1.0, value=float(case["pct_spliced"]), step=key_value(source.sources, "validation_tolerance") or 0.01)
                confinement = st.number_input("Factor de confinamiento", min_value=key_value(source.sources, "zero") or 0.0, value=key_value(source.sources, "one") or 0.0, step=key_value(source.sources, "one") or 1.0)

        values = {
            "fc": fc,
            "fy": fy,
            "db": db,
            "cover": cover,
            "spacing": spacing,
            "psi_t": psi_t,
            "psi_e": psi_e,
            "psi_s": psi_s,
            "lambda": lambda_factor,
            "ktr": ktr,
            "as_req": as_req,
            "as_prov": as_prov,
            "pct_spliced": pct_spliced,
            "confinement": confinement,
        }
        hook = calculate_hook(source, hook_index, db, bar_type, range_label)
        bend = calculate_bend(source, db, bar_type, range_label)
        development = calculate_development_and_laps(source, values)

        st.subheader("Dimensiones y resultados trazables")
        dim1, dim2, dim3, dim4 = st.columns(4)
        with dim1:
            render_source_card("Diámetro comercial seleccionado", selected_bar["commercial"], "", selected_bar["reference"], selected_bar["state"], f"Dato directo de BARRAS!A{selected_bar['row']}")
        with dim2:
            render_source_card("db de BARRAS", selected_bar["db"], "mm", selected_bar["reference"], selected_bar["state"], selected_bar["db_formula"])
        with dim3:
            render_source_card("Área de BARRAS", selected_bar["area"], "mm²", selected_bar["reference"], selected_bar["state"], selected_bar["area_formula"])
        with dim4:
            render_source_card("Perímetro de BARRAS", selected_bar["perimeter"], "mm", selected_bar["reference"], selected_bar["state"], selected_bar["perimeter_formula"])

        row1, row2, row3 = st.columns(3)
        with row1:
            render_trace_card(hook, source)
        with row2:
            render_trace_card(bend, source)
        with row3:
            render_trace_card(development["ld_tension"], source)

        row4, row5, row6 = st.columns(3)
        with row4:
            render_trace_card(development["ld_compression"], source)
        with row5:
            render_trace_card(development["ldg"], source)
        with row6:
            render_trace_card(development["lap_tension"], source)

        row7, row8 = st.columns(2)
        with row7:
            render_trace_card(development["lap_compression"], source)
        with row8:
            render_trace_card(development["column_splice"], source)

        st.subheader("Fuente y trazabilidad")
        st.dataframe(
            [
                {
                    "dato": item["key"],
                    "capa": item["layer"],
                    "artículo o tabla": item["reference"],
                    "página": item["page"],
                    "estado": item["state"],
                    "valor": item["value"],
                    "unidad": item["unit"],
                }
                for item in source.sources.values()
            ],
            use_container_width=True,
            hide_index=True,
        )

    with details:
        st.subheader("Lámina técnica de acero de refuerzo")
        if PLATE_PATH.exists():
            st.image(str(PLATE_PATH), caption="Lámina técnica — detalle de acero de refuerzo")
        else:
            st.error(f"Lámina NO VERIFICADA: no se encontró {PLATE_PATH}")
        st.caption("La lámina es material educativo; las dimensiones deben verificarse contra el Excel y la Norma E.060 vigente.")


if __name__ == "__main__":
    main()
